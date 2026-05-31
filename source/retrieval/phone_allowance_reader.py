from pathlib import Path

_PhoneAllowance_DF = None


def get_phone_allowance_table() -> str:
    """
    Đọc bảng phụ cấp điện thoại và trả về dưới dạng Markdown table.
    Bảng có cấu trúc: Level (dòng) × Department (cột) với giá trị phụ cấp (nghìn VNĐ).
    """
    global _PhoneAllowance_DF

    if _PhoneAllowance_DF is not None:
        return _PhoneAllowance_DF

    from openpyxl import load_workbook

    path = Path("raw-data/Chi tiết Quy định phụ cấp điện thoại.xlsx")
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 5:
        return "Không có dữ liệu phụ cấp điện thoại."

    # Row 2 (index 1): Department names - forward fill merged cells
    raw_depts = [str(val).strip() if val is not None else "" for val in rows[1][1:]]
    departments = []
    last_dept = ""
    for d in raw_depts:
        if d and d != "None":
            last_dept = d
        departments.append(last_dept)

    # Row 3 (index 2): Group names (sub-departments)
    raw_groups = [str(val).strip() if val is not None else "" for val in rows[2][1:]]

    # Build header: Department + Group combined
    headers = []
    for dept, grp in zip(departments, raw_groups):
        if grp and grp != "None":
            headers.append(f"{dept} - {grp}")
        else:
            headers.append(dept)

    # Data rows (from row 5 onwards, col 0 = Level, cols 1+ = amounts)
    lines = []
    header_line = "| Level | " + " | ".join(headers) + " |"
    separator = "|---|" + "|".join(["---:"] * len(headers)) + "|"
    lines.append(header_line)
    lines.append(separator)

    for row in rows[4:]:
        if row[0] is None:
            continue
        level = str(row[0]).strip()
        values = []
        for val in row[1:]:
            if val is not None:
                values.append(f"{int(val)}K")
            else:
                values.append("-")
        lines.append(f"| {level} | " + " | ".join(values) + " |")

    result = "\n".join(lines)
    _PhoneAllowance_DF = result
    return result
